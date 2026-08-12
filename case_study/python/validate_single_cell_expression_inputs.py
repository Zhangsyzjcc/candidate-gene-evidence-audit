#!/usr/bin/env python3
"""Validate registered single-cell expression inputs after protocol freeze.
Streams files from tar archives and writes only derived CSV metadata/results."""
import csv, gzip, math, re, tarfile
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; DATE='2026-08-01'; stats=ROOT/'results/statistics'; interim=ROOT/'data/interim/GEO/single_cell'; stats.mkdir(parents=True,exist_ok=True); interim.mkdir(parents=True,exist_ok=True)
validation=[]
def numeric_summary(vals):
 x=[]
 for v in vals:
  try: x.append(float(v))
  except ValueError: pass
 return dict(lrrk2_nonzero_cells=sum(v>0 for v in x),lrrk2_min=min(x) if x else '',lrrk2_max=max(x) if x else '',lrrk2_mean=sum(x)/len(x) if x else '')
for ds in ('GSE131928','GSE103224'):
 with tarfile.open(ROOT/f'data/raw/GEO/single_cell/{ds}/{ds}_RAW.tar') as t:
  for m in t.getmembers():
   if not m.name.endswith('.gz'): continue
   with gzip.GzipFile(fileobj=t.extractfile(m)) as f:
    first=f.readline().decode('utf-8','replace').rstrip('\r\n').split('\t')
    has_header=(ds=='GSE131928'); id_cols=1 if has_header else 2
    n_cells=max(0,len(first)-id_cols); genes=0; match=[]; integer_probe=True; probe=0
    if not has_header:
     fields=first; genes=1
     for v in fields[id_cols:id_cols+100]:
      try: integer_probe &= float(v).is_integer()
      except ValueError: integer_probe=False
     probe=1
     if any(v.upper()=='LRRK2' for v in fields[:id_cols]): match=fields[id_cols:]
    for raw in f:
      fields=raw.decode('utf-8','replace').rstrip('\r\n').split('\t'); genes+=1
      if probe<100:
       for v in fields[id_cols:id_cols+100]:
        try: integer_probe &= float(v).is_integer()
        except ValueError: integer_probe=False
       probe+=1
      if any(v.upper()=='LRRK2' for v in fields[:id_cols]): match=fields[id_cols:]
    scale='TPM_processed' if 'TPM' in m.name else ('integer_like_filtered_matrix' if integer_probe else 'noninteger_filtered_matrix')
    z=numeric_summary(match)
    validation.append(dict(dataset=ds,gsm=re.match(r'(GSM\d+)',m.name).group(1),sample_id=re.sub(r'^GSM\d+_','',m.name).split('.')[0],assay='scRNA',input_format='TPM_matrix' if 'TPM' in m.name else 'filtered_text_matrix',n_features=genes,n_cells=n_cells,value_scale=scale,integer_probe_first_100_genes=integer_probe,lrrk2_feature_matches=1 if match else 0,**z))
ds='GSE138794'
with tarfile.open(ROOT/f'data/raw/GEO/single_cell/{ds}/{ds}_RAW.tar') as t:
 groups={}
 for m in t.getmembers():
  gsm=re.match(r'(GSM\d+)',m.name)
  if gsm: groups.setdefault(gsm.group(1),[]).append(m)
 for gsm,mem in sorted(groups.items()):
  names=';'.join(x.name for x in mem)
  if 'ATAC' in names or 'peaks' in names: continue
  feat=next((x for x in mem if 'features' in x.name or 'genes' in x.name),None); bar=next((x for x in mem if 'barcodes' in x.name),None); mat=next((x for x in mem if 'matrix.mtx' in x.name),None)
  if not (feat and bar and mat): continue
  with gzip.GzipFile(fileobj=t.extractfile(feat)) as f:
   feature_rows=[line.decode('utf-8','replace').rstrip('\r\n').split('\t') for line in f]
  with gzip.GzipFile(fileobj=t.extractfile(bar)) as f: nbar=sum(1 for _ in f)
  with gzip.GzipFile(fileobj=t.extractfile(mat)) as f:
   line=f.readline().decode(); integer='integer' in line
   line=f.readline().decode()
   while line.startswith('%'): line=f.readline().decode()
   dims=[int(x) for x in line.split()[:3]]
  matches=[i+1 for i,x in enumerate(feature_rows) if any(v.upper()=='LRRK2' for v in x)]
  validation.append(dict(dataset=ds,gsm=gsm,sample_id=re.sub(r'^GSM\d+_','',mem[0].name).split('_')[0],assay='snRNA' if 'sn_' in names.lower() else 'scRNA',input_format='10x_mtx_triplet',n_features=dims[0],n_cells=dims[1],value_scale='integer_counts' if integer else 'matrix_market_noninteger',integer_probe_first_100_genes=integer,lrrk2_feature_matches=len(matches),lrrk2_nonzero_cells='not_scanned_sparse_matrix',lrrk2_min='',lrrk2_max='',lrrk2_mean=''))
out=stats/f'single_cell_expression_input_validation_{DATE}.csv'
with out.open('w',encoding='utf-8',newline='') as f: w=csv.DictWriter(f,fieldnames=validation[0].keys()); w.writeheader(); w.writerows(validation)
# GSE131928 cell-to-patient metadata from the registered workbook.
sheet=load_workbook(ROOT/'data/raw/GEO/single_cell/GSE131928/GSE131928_single_cells_tumor_name_and_adult_or_peidatric.xlsx',read_only=True,data_only=True).active
rows=list(sheet.iter_rows(values_only=True)); header=[str(x).strip() if x is not None else '' for x in rows[43][:9]]; meta=[]
for row in rows[44:]:
 if not row[0]: continue
 rec={header[i]:row[i] for i in range(9)}; meta.append(rec)
p=interim/f'GSE131928_cell_patient_metadata_{DATE}.csv'; p.parent.mkdir(parents=True,exist_ok=True)
with p.open('w',encoding='utf-8',newline='') as f: w=csv.DictWriter(f,fieldnames=header); w.writeheader(); w.writerows(meta)
# GSE138794 supplied cell-type labels, preserving submitter labels verbatim.
ct=[]
for pth in (ROOT/'data/raw/GEO/single_cell/GSE138794').glob('*cell_types.txt.gz'):
 assay='snRNA' if 'snRNA' in pth.name else 'scRNA'
 with gzip.open(pth,'rt',encoding='utf-8') as f:
  for line in f:
   cell,label=line.rstrip('\r\n').split(maxsplit=1); ct.append(dict(assay=assay,cell_id=cell,sample_id=cell.split('_')[0],submitter_cell_type=label))
p=interim/f'GSE138794_submitter_cell_types_{DATE}.csv'
with p.open('w',encoding='utf-8',newline='') as f: w=csv.DictWriter(f,fieldnames=ct[0].keys()); w.writeheader(); w.writerows(ct)
print(f'Validated {len(validation)} RNA matrices; exported {len(meta)} and {len(ct)} cell metadata rows.')
